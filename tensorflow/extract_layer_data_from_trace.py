#!/usr/bin/env python3

import os
import sys
import re
import json

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

KEY_TRACE_EVENTS = "traceEvents"
KEY_CATEGORY = "cat"
KEY_ARGS = "args"
KEY_NAME = "name"
KEY_PID = "pid"
KEY_OP = "op"
KEY_KERNEL = "kernel"
KEY_DURATION = "dur"


VALUE_PROCESS_NAME = "process_name"
VALUE_CATEGORY_DATAFLOW = "DataFlow"
VALUE_CATEGORY_OP = "Op"


def print_data(data):
  total_duration = 0
  for scope, kernels in data.items():
    for kernel, duration in kernels:
      total_duration += duration
  print (total_duration)


def determine_pid(trace_events, process_name):
  pid = -1
  for event in trace_events:
    if KEY_NAME not in event.keys():
      continue
    if event[KEY_NAME] != VALUE_PROCESS_NAME:
      continue
    if event[KEY_ARGS][KEY_NAME] == process_name:
      pid = event[KEY_PID]
      break
  return pid


def extract_op_data(trace_events, op_to_extract):

  gpu_stream_all_pid = determine_pid(trace_events, "/device:GPU:0/stream:all Compute")

  data = {}
  for event in trace_events:
    if KEY_CATEGORY not in event.keys():
      continue
    if event[KEY_CATEGORY] != VALUE_CATEGORY_OP:
      continue
    if event[KEY_PID] != gpu_stream_all_pid:
      continue
    op_name = event[KEY_ARGS][KEY_OP]
    if op_name == op_to_extract:
      scope = event[KEY_ARGS][KEY_NAME]
      kernel = event[KEY_ARGS].get(KEY_KERNEL, "<no_kernel_name>")
      duration = event[KEY_DURATION]
      scope_kernels = data.get(scope, [])
      scope_kernels.append((kernel, duration))
      data[scope] = scope_kernels

  return data


def generate_comparison_data_xlsx(op_name, mi100_trace_events, v100_trace_events, scope_to_config_map, xlsx_file):
  mi100_data = extract_op_data(mi100_trace_events, op_name)
  v100_data = extract_op_data(v100_trace_events, op_name)
    
  mi100_scopes = sorted(mi100_data.keys())
  v100_scopes = sorted(v100_data.keys())
  if mi100_scopes != v100_scopes:
    print("Data collection error : scopes do not match")

  workbook = Workbook()
  sheet = workbook.active

  dark_green_font = Font(bold=True, color='008800')
  light_green_fill = PatternFill('solid', fgColor='DDFFDD')
  dark_red_font = Font(bold=True, color='880000')
  light_red_fill = PatternFill('solid', fgColor='FFDDDD')
  
  right_align = Alignment(horizontal="right")

  header = ["Model Scope and Config",
            "MI100(us)", "V100(us)", "V100/MI100",
            "MI100 Kernel Name", "V100 Kernel Name"]
  sheet.append(header)
  for i in range(len(header)):
    sheet.cell(row=sheet._current_row, column=i+1).font = Font(bold=True)
  sheet.append([])
  
  for scope in mi100_scopes:

    config = scope_to_config_map[scope]
    
    mi100_kernels = mi100_data[scope]
    v100_kernels = v100_data[scope]

    num_mi100_kernels = len(mi100_kernels)
    num_v100_kernels = len(v100_kernels)
      
    num_kernels = max(2, num_mi100_kernels, num_v100_kernels)
    mi100_duration = 0
    v100_duration = 0
    for i in range(num_kernels):
      mi100_kernel_name, mi100_kernel_duration = mi100_kernels[i] if i < num_mi100_kernels else (" ","")
      v100_kernel_name, v100_kernel_duration = v100_kernels[i] if i < num_v100_kernels else (" ","")
      
      line = [""]
      if i == 0: line = [scope]
      if i == 1: line = [config]
      line.append(mi100_kernel_duration)
      line.append(v100_kernel_duration)
      line.append("")
      line.append(mi100_kernel_name)
      line.append(v100_kernel_name)
      sheet.append(line)
      
      mi100_duration += mi100_kernel_duration if mi100_kernel_duration else 0
      v100_duration += v100_kernel_duration if v100_kernel_duration else 0

    v100_mi100_ratio = v100_duration / mi100_duration
    line = ["layer total"]
    line.append(mi100_duration)
    line.append(v100_duration)
    line.append(v100_mi100_ratio)
    sheet.append(line)

    sheet.cell(row=sheet._current_row, column=1).alignment = right_align
    sheet.cell(row=sheet._current_row, column=4).font = dark_red_font if v100_mi100_ratio < 1.0 else dark_green_font
    sheet.cell(row=sheet._current_row, column=4).fill = light_red_fill if v100_mi100_ratio < 1.0 else light_green_fill

    sheet.append([])

  sheet.append([])
  line = []
  line.append("Total Kernel Time (us)")
  line.append("=SUM(B2:B{})/2".format(sheet._current_row - 1))
  line.append("=SUM(C2:C{})/2".format(sheet._current_row - 1))
  sheet.append(line)
  
  sheet.column_dimensions['A'].width = 90
  sheet.column_dimensions['B'].width = 15
  sheet.column_dimensions['C'].width = 15
  sheet.column_dimensions['D'].width = 15
  sheet.column_dimensions['E'].width = 50
  sheet.column_dimensions['F'].width = 50
    
  workbook.save(filename=xlsx_file)


def get_trace_events(trace_file):
  with open(trace_file) as f:
    json_data = json.load(f)
  return json_data[KEY_TRACE_EVENTS]
  

if __name__ == '__main__':

  mi100_trace_events = get_trace_events("mi100.json")
  v100_trace_events = get_trace_events("v100.json")
  with open("scope_to_config.json") as f:
    scope_to_config_map = json.load(f)
    
  generate_comparison_data_xlsx("Conv2D", mi100_trace_events, v100_trace_events, scope_to_config_map, "Conv2D_compare.xlsx")
  generate_comparison_data_xlsx("Conv2DBackpropFilter", mi100_trace_events, v100_trace_events, scope_to_config_map, "Conv2DBackpropFilter_compare.xlsx")
  generate_comparison_data_xlsx("Conv2DBackpropInput", mi100_trace_events, v100_trace_events, scope_to_config_map, "Conv2DBackpropInput_compare.xlsx")
