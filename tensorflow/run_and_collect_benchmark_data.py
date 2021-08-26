#!/usr/bin/env python3

import argparse
import json
import os
import re
import subprocess
import sys

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font

os.environ["HIP_VISIBLE_DEVICES"]="0";

KEY_OPTIONS = "options"
KEY_ENV_VARS = "env_vars"

def run_shell_command(cmd, env_vars):

    all_env_vars = " ".join(["{}={}".format(key, value) for key, value in env_vars.items()])
    shell_cmd = " ".join(cmd)
    full_cmd = " ".join([all_env_vars, shell_cmd])
    print(full_cmd)
    result = subprocess.run(full_cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True)
    if result.returncode != 0:
        print("FAILED - {}".format(" ".join(cmd)))
        print("         {}".format(result.stdout.decode()))
        print("         {}".format(result.stderr.decode()))
    return result


def run_tf_cnn_benchmark(model, options, env_vars):

    run_benchmark_cmd = [
        "python3",
        "scripts/tf_cnn_benchmarks/tf_cnn_benchmarks.py",
        "--model={}".format(model),
    ]
    for option in options:
        run_benchmark_cmd.append(option)

    result = run_shell_command(run_benchmark_cmd, env_vars)
    return (result.returncode, result.stdout.decode(), result.stderr.decode())



def collect_tf_cnn_benchmark_perf_data():

    models = [
        "alexnet",
        "googlenet",
        "inception3",
        "inception4",
        "lenet",
        # "mobilenet",
        # "nasnet",
        # "nasnetlarge",
        # "ncf",
        # "overfeat",
        "resnet50",
        "resnet50_v1.5",
        "resnet101",
        "resnet152_v2",
        # "trivial",
        # "vgg11",
        "vgg16",
        "vgg19",
    ]

    configs = {
        # "TRADITIONAL BACKEND - FP32" : {
        #     KEY_OPTIONS : [],
        #     KEY_ENV_VARS : {"TF_ROCM_FUSION_ENABLE":"1",}
        #     },

        "TRADITIONAL BACKEND - FP16" : {
            KEY_OPTIONS : ["--use_fp16"],
            KEY_ENV_VARS : {"TF_ROCM_FUSION_ENABLE":"1", "MIOPEN_FIND_MODE":"1"}
            },

        # "XLA BACKEND - FP32" : {
        #     KEY_OPTIONS : ["--xla_compile"],
        #     KEY_ENV_VARS : {"TF_ROCM_FMA_DISABLE":"1",}
        #     },
        
        # "XLA BACKEND - FP16" : {
        #     KEY_OPTIONS : ["--use_fp16", "--xla_compile"],
        #     KEY_ENV_VARS : {"TF_ROCM_FMA_DISABLE":"1",}
        #     },
        # "AMP" : {
        #     KEY_OPTIONS : ["--auto_mixed_precision"],
        #     KEY_ENV_VARS : {"TF_ROCM_FMA_DISABLE":"1",}
        #     },

        # "baseline" : {
        #     KEY_OPTIONS : [],
        #     KEY_ENV_VARS : {"TF_ROCM_FMA_DISABLE":"1",}
        #     },

        # "AMP + FUSION" : {
        #     KEY_OPTIONS : ["--auto_mixed_precision"],
        #     KEY_ENV_VARS : {"TF_ROCM_FUSION_ENABLE":"1",}
        #     },

        # "AMP + XLA" : {
        #     KEY_OPTIONS : ["--auto_mixed_precision", "--xla"],
        #     KEY_ENV_VARS : {"TF_ROCM_FMA_DISABLE":"1",}
        #     },

        }

    def add_option_to_all_configs(option):
        for key, value in configs.items():
            value[KEY_OPTIONS].append(option)

    # add_option_to_all_configs("--batch_size=256")

    def display_configs():
        for key, value in configs.items():
            print("{}:".format(key))
            print("\toptions : ", value[KEY_OPTIONS])
            print("\tenv_vars : ", value[KEY_ENV_VARS])

    display_configs()
                  
    N = 5
    
    data = {}
    data["models"] = models
    data["configs"] = configs
    data["N"] = N
    
    def run_N_times(N, model, options=[], env_vars = {}):
        config_data = {}
        for i in range(N):
            run_data = run_tf_cnn_benchmark(model, options, env_vars)
            config_data["run_{}".format(i)] = run_data
        return config_data

    for model in models:
        for name, config in configs.items():
            config_data = run_N_times(N, model, config[KEY_OPTIONS], config[KEY_ENV_VARS])
            data["{}_{}".format(model, name)] = config_data

    return data


def gather_stats(data):

    models = data["models"]
    configs = data["configs"]
    N = data["N"]
    
    stats = {}
    stats["models"] = models
    stats["configs"] = configs
    stats["N"] = N

    def get_stats(N, config_data):
        run_stats = []
        for i in range(N):
            returncode, stdout, stderr = config_data["run_{}".format(i)]
            if returncode == 0:
                pattern = r'total images/sec: ([0-9\.]+)'
                match = re.search(pattern, stdout)
                assert(match is not None)
                run_stats.append(float(match.group(1)))
            else :
                run_stats.append("error")
        return run_stats
    
    for model in models:
        model_stats = {}
        for name, config in configs.items():
            config_data = data["{}_{}".format(model, name)]
            model_stats[name] = get_stats(N, config_data)
        stats[model] = model_stats

    return stats


def dump_stats(excel_file, stats):

    workbook = Workbook()
    sheet = workbook.active

    workbook = Workbook()
    sheet = workbook.active

    models = stats["models"]
    configs = stats["configs"]
    N = stats["N"]
    
    header_1 = ["", ""]
    for name, config in configs.items():
        header_1.append(name)
        for i in range(N):
            header_1.append("")
        header_1.append("")
    print (header_1)
    sheet.append(header_1)
    
    header_2 = ["benchmark", ""]
    for name, config in configs.items():
        for i in range(N):
            header_2.append("run_{}".format(i))
        header_2.append("AVERAGE") 
        header_2.append("") 
    print (header_2)
    sheet.append(header_2)

    for model in models:
        row = [model, ""]
        model_stats = stats[model]
        for name, config in configs.items():
            config_stats = model_stats[name]
            for i in range(N):
                row.append(config_stats[i])
            row.append("")
            row.append("")
        print(row)
        sheet.append(row)
            
    workbook.save(filename=excel_file)


def dump_stats_transposed(excel_file, stats):

    workbook = Workbook()
    sheet = workbook.active

    workbook = Workbook()
    sheet = workbook.active

    models = stats["models"]
    configs = stats["configs"]
    N = stats["N"]
    
    header_1 = ["", ""]
    for name, config in configs.items():
        header_1.append(name)
        for i in range(N):
            header_1.append("")
        header_1.append("")
    print (header_1)
    sheet.append(header_1)
    
    header_2 = ["benchmark", ""]
    for name, config in configs.items():
        for i in range(N):
            header_2.append("run_{}".format(i))
        header_2.append("AVERAGE") 
        header_2.append("") 
    print (header_2)
    sheet.append(header_2)

    for model in models:
        row = [model, ""]
        model_stats = stats[model]
        for name, config in configs.items():
            config_stats = model_stats[name]
            for i in range(N):
                row.append(config_stats[i])
            row.append("")
            row.append("")
        print(row)
        sheet.append(row)
            
    workbook.save(filename=excel_file)


if __name__ == '__main__':
           
    parser = argparse.ArgumentParser()
    parser.add_argument("-d", "--dump_stats_from", default=None)
    args = parser.parse_args()

    if (args.dump_stats_from):
        filename = args.dump_stats_from
        with open(filename) as f:
            data = json.loads(f.read())
            stats = gather_stats(data)
            dump_stats(filename.replace("json", "xlsx"), stats)
            # dump_stats_transposed(filename.replace("json", "xlsx"), stats)

    else :
        data = collect_tf_cnn_benchmark_perf_data()
        now = datetime.now()
        filename = "bm_data_{}.json".format(now.strftime("%y%m%d_%H%M"))
        with open(filename, "w") as f:
            f.write(json.dumps(data))
